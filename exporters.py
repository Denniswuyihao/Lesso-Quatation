from __future__ import annotations

import base64
import html
import mimetypes
import os
from datetime import datetime
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from typing import Optional

import pandas as pd
import requests
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Image as PDFImage
from reportlab.platypus import KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BASE_DIR = Path(__file__).resolve().parent
FONT_DIR = BASE_DIR / "assets" / "fonts"

BRAND_BLUE = "#1F4E78"
BRAND_DARK = "#17365D"
BORDER_BLUE = "#D9E2F3"
LIGHT_BG = "#F7F9FC"
SOFT_BG = "#FAFAFA"
TEXT_DARK = "#222222"
TEXT_MUTED = "#666666"


# -------------------------
# Common labels
# -------------------------

def _headers(lang: str) -> list[str]:
    if lang == "en":
        return [
            "SAP", "Category", "Product Name", "Model", "Image", "Description", "Base Price (USD)", "Quote Price (USD)",
            "Qty", "Amount (USD)", "Stock", "CBM/PC", "Total CBM", "Package"
        ]
    return [
        "SAP号", "分类", "品名", "型号", "图片", "描述", "原价(USD)", "报价(USD)",
        "数量", "金额(USD)", "库存", "包装体积/件", "总CBM", "包装"
    ]

def _title(lang: str) -> str:
    return "LESSO Plumbing & Sanitary Ware Quotation" if lang == "en" else "LESSO 联塑水暖卫浴外贸报价单"


def _field_labels(lang: str) -> dict[str, str]:
    if lang == "en":
        return {
            "customer": "Customer", "quote_no": "Quote No.", "date": "Date", "rule": "Price Rule",
            "total": "Total", "prepared_by": "Prepared By", "currency": "Currency", "page": "Page",
            "sku": "SKU", "qty": "Total Qty", "amount": "Total Amount (USD)", "cbm": "Total CBM",
            "terms": "Remarks", "validity": "Validity", "signature": "Confirmation"
        }
    return {
        "customer": "客户", "quote_no": "报价单号", "date": "日期", "rule": "价格规则",
        "total": "合计", "prepared_by": "制单人", "currency": "币种", "page": "页码",
        "sku": "SKU数量", "qty": "总数量", "amount": "合计金额(USD)", "cbm": "总体积",
        "terms": "备注", "validity": "有效期", "signature": "确认"
    }


def _name(row: pd.Series | dict, lang: str) -> str:
    """For English exports, always prefer English product name."""
    if lang == "en":
        return str(row.get("en_name") or row.get("cn_name") or "")
    return str(row.get("cn_name") or row.get("en_name") or "")


def _safe_text(value: object) -> str:
    return html.escape(str(value or ""), quote=False)


# -------------------------
# Images
# -------------------------

def resolve_image_url(image_url: str | None, sap: str | None = None) -> str:
    """Return the cloud image URL. Stable version intentionally does not use local images."""
    value = str(image_url or "").strip()
    if value.startswith("http://") or value.startswith("https://"):
        return value
    return ""


def image_to_data_uri(image_url: str | None, sap: str | None = None) -> str:
    """Return a URL for Streamlit ImageColumn preview. Empty string means no image."""
    return resolve_image_url(image_url, sap)


@lru_cache(maxsize=512)
def _download_image_bytes(image_url: str | None, timeout: int = 12) -> Optional[bytes]:
    url = resolve_image_url(image_url)
    if not url:
        return None
    try:
        resp = requests.get(url, timeout=timeout)
        resp.raise_for_status()
        content_type = resp.headers.get("content-type", "")
        if content_type and not content_type.lower().startswith("image/"):
            return None
        return resp.content
    except Exception:
        return None


def _resize_for_excel_bytes(data: bytes, max_px: int = 82) -> tuple[int, int]:
    with PILImage.open(BytesIO(data)) as im:
        w, h = im.size
    ratio = min(max_px / max(w, 1), max_px / max(h, 1), 1)
    return int(w * ratio), int(h * ratio)


def _excel_image_from_url(image_url: str | None, max_px: int = 82) -> tuple[Optional[XLImage], Optional[BytesIO]]:
    data = _download_image_bytes(image_url)
    if not data:
        return None, None
    bio = BytesIO(data)
    try:
        xl_img = XLImage(bio)
        xl_img.width, xl_img.height = _resize_for_excel_bytes(data, max_px=max_px)
        return xl_img, bio
    except Exception:
        return None, None


def _pdf_image_from_url(image_url: str | None, max_w: float = 21 * mm, max_h: float = 18 * mm) -> Optional[PDFImage]:
    data = _download_image_bytes(image_url)
    if not data:
        return None
    try:
        with PILImage.open(BytesIO(data)) as im:
            w, h = im.size
        ratio = min(max_w / max(w, 1), max_h / max(h, 1))
        return PDFImage(BytesIO(data), width=w * ratio, height=h * ratio)
    except Exception:
        return None


# Backward-compatible name. It now intentionally ignores local files.
def resolve_image_path(image_path: str | None, sap: str | None = None) -> None:
    return None


# -------------------------
# Chinese PDF font handling
# -------------------------

_FONT_CACHE: dict[str, str] = {}


def _candidate_chinese_fonts() -> list[Path]:
    """
    Return possible embeddable Chinese font paths.

    Best practice for searchable Chinese PDF:
    1. Put a .ttf/.otf font in assets/fonts/ and set CHINESE_FONT_PATH if needed.
    2. Otherwise the program tries common system fonts.

    We do NOT bundle font files in this project. This avoids font license issues.
    """
    candidates: list[Path] = []

    env_path = os.getenv("CHINESE_FONT_PATH")
    if env_path:
        candidates.append(Path(env_path).expanduser())

    if FONT_DIR.exists():
        for ext in ("*.ttf", "*.otf"):
            candidates.extend(sorted(FONT_DIR.glob(ext)))

    candidates.extend([
        # Windows
        Path("C:/Windows/Fonts/msyh.ttf"),
        Path("C:/Windows/Fonts/simhei.ttf"),
        Path("C:/Windows/Fonts/simsun.ttf"),
        # macOS
        Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
        Path("/Library/Fonts/Arial Unicode.ttf"),
        Path("/Library/Fonts/Microsoft/Microsoft Yahei.ttf"),
        Path("/Library/Fonts/Microsoft/SimHei.ttf"),
        # Linux / common server fonts
        Path("/usr/share/fonts/truetype/arphic-gbsn00lp/gbsn00lp.ttf"),
        Path("/usr/share/fonts/truetype/arphic-gkai00mp/gkai00mp.ttf"),
    ])

    seen = set()
    unique: list[Path] = []
    for p in candidates:
        key = str(p)
        if key not in seen:
            seen.add(key)
            unique.append(p)
    return unique


def _register_ttf_font(font_path: Path) -> Optional[str]:
    """Register an embeddable TrueType/OpenType font. Return ReportLab font name if successful."""
    if not font_path.exists() or not font_path.is_file():
        return None

    cache_key = str(font_path.resolve())
    if cache_key in _FONT_CACHE:
        return _FONT_CACHE[cache_key]

    font_name = "QuoteCN_" + str(abs(hash(cache_key)))
    try:
        pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
        _FONT_CACHE[cache_key] = font_name
        return font_name
    except Exception:
        return None


def get_pdf_font_status() -> str:
    """Human-readable status shown in the Streamlit sidebar."""
    for path in _candidate_chinese_fonts():
        font_name = _register_ttf_font(path)
        if font_name:
            return f"中文PDF字体：已嵌入 {path.name}"
    return "中文PDF字体：使用备用中文CID字体；如需可复制/搜索中文，请在 assets/fonts/ 放入中文TTF/OTF字体。"


def _register_cn_font() -> str:
    for path in _candidate_chinese_fonts():
        font_name = _register_ttf_font(path)
        if font_name:
            return font_name

    fallback = "STSong-Light"
    try:
        pdfmetrics.registerFont(UnicodeCIDFont(fallback))
    except Exception:
        pass
    return fallback


def _pdf_para(text_value: object, style: ParagraphStyle) -> Paragraph:
    return Paragraph(_safe_text(text_value), style)


def _pdf_lines(lines: list[object], style: ParagraphStyle) -> Paragraph:
    """Safe multi-line paragraph. User text is escaped; only our own <br/> tags are kept."""
    clean = [_safe_text(x) for x in lines if str(x or "").strip()]
    return Paragraph("<br/>".join(clean), style)


def _contains_cjk(value: object) -> bool:
    text = str(value or "")
    return any("\u4e00" <= ch <= "\u9fff" for ch in text)


def _pdf_para_auto(text_value: object, styles: dict[str, ParagraphStyle], style_key: str) -> Paragraph:
    style = styles.get(style_key + "_cjk") if _contains_cjk(text_value) else styles.get(style_key)
    return _pdf_para(text_value, style or styles[style_key])


def _pdf_lines_auto(lines: list[object], styles: dict[str, ParagraphStyle], style_key: str) -> Paragraph:
    style = styles.get(style_key + "_cjk") if any(_contains_cjk(x) for x in lines) else styles.get(style_key)
    return _pdf_lines(lines, style or styles[style_key])


# -------------------------
# Excel export
# -------------------------

def export_quote_excel(items: pd.DataFrame, customer: str, quote_no: str, price_note: str = "", lang: str = "zh") -> bytes:
    """Export an A4 portrait Excel quotation. lang='zh' or 'en'."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation" if lang == "en" else "报价单"

    labels = _field_labels(lang)
    headers = _headers(lang)
    max_col = len(headers)

    navy = "1F4E78"
    light_fill = "F7F9FC"
    border_color = "D9E2F3"
    total_fill = "EAF2F8"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(1, 1).value = _title(lang)
    ws.cell(1, 1).font = Font(name="Microsoft YaHei", size=17, bold=True, color="FFFFFF")
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=navy)
    ws.row_dimensions[1].height = 32

    info_fill = PatternFill("solid", fgColor=light_fill)
    ws["A2"] = labels["customer"]
    ws["B2"] = customer or ""
    ws["D2"] = labels["quote_no"]
    ws["E2"] = quote_no
    ws["G2"] = labels["date"]
    ws["H2"] = datetime.now().strftime("%Y-%m-%d")
    ws["J2"] = labels["rule"]
    ws["K2"] = price_note
    for cell in ws[2]:
        cell.fill = info_fill
        cell.alignment = Alignment(vertical="center")
        cell.font = Font(name="Microsoft YaHei", size=9)
    ws.row_dimensions[2].height = 24

    start_row = 4
    for col_idx, header in enumerate(headers, 1):
        ws.cell(start_row, col_idx).value = header

    header_font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=9)
    thin = Side(style="thin", color=border_color)
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for cell in ws[start_row]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[start_row].height = 26

    for _, row in items.iterrows():
        ws.append([
            row.get("sap", ""), row.get("category", ""), _name(row, lang), row.get("model", ""), "",
            row.get("description", ""), row.get("base_price", 0), row.get("quote_price", 0), row.get("quantity", 1),
            row.get("amount", 0), row.get("stock", 0), row.get("packing_volume", 0), row.get("total_volume", 0),
            row.get("package_info", ""),
        ])
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = 68
        image_url = row.get("image_url")
        if image_url:
            try:
                xl_img, img_buffer = _excel_image_from_url(str(image_url))
                if xl_img is not None:
                    # Keep the BytesIO object alive until workbook.save().
                    if not hasattr(wb, "_quote_image_buffers"):
                        wb._quote_image_buffers = []
                    wb._quote_image_buffers.append(img_buffer)
                    ws.add_image(xl_img, f"E{current_row}")
                else:
                    ws[f"E{current_row}"] = "Image error" if lang == "en" else "图片读取失败"
            except Exception:
                ws[f"E{current_row}"] = "Image error" if lang == "en" else "图片读取失败"

    last_row = ws.max_row
    total_row = last_row + 1

    total_label = "TOTAL" if lang == "en" else "合计"
    # 总计行紧贴产品表，不留空行；左侧合并，金额放在“金额(USD)”列下方，体积放在“总CBM”列下方。
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
    ws.cell(total_row, 1).value = total_label
    ws.cell(total_row, 9).value = f"=SUM(I{start_row + 1}:I{last_row})"
    ws.cell(total_row, 10).value = f"=SUM(J{start_row + 1}:J{last_row})"
    ws.cell(total_row, 13).value = f"=SUM(M{start_row + 1}:M{last_row})"

    for col_idx in range(1, max_col + 1):
        cell = ws.cell(total_row, col_idx)
        cell.fill = PatternFill("solid", fgColor=total_fill)
        cell.border = border
        cell.font = Font(name="Microsoft YaHei", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(total_row, 10).number_format = '"USD" #,##0.00'
    ws.cell(total_row, 13).number_format = "0.0000"
    ws.row_dimensions[total_row].height = 26

    widths = {
        "A": 12, "B": 11, "C": 24, "D": 14, "E": 13, "F": 28, "G": 11, "H": 11,
        "I": 7, "J": 12, "K": 8, "L": 11, "M": 9, "N": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=start_row, max_row=last_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            if cell.row > start_row:
                cell.font = Font(name="Microsoft YaHei", size=9)
    for row_idx in range(start_row + 1, last_row + 1):
        ws[f"C{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"F{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"N{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for col in ["G", "H", "J"]:
            ws[f"{col}{row_idx}"].number_format = '"USD" #,##0.00'
        for col in ["L", "M"]:
            ws[f"{col}{row_idx}"].number_format = "0.0000"

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{start_row}:N{last_row}"

    # A4 portrait print settings
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.print_title_rows = f"{start_row}:{start_row}"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# -------------------------
# PDF export
# -------------------------

def _make_pdf_styles(lang: str) -> dict[str, ParagraphStyle]:
    cn_font = _register_cn_font()
    # English exports use Helvetica for a cleaner Western business look.
    # Extra CJK fallback styles are used only when source data still contains Chinese.
    base_font = "Helvetica" if lang == "en" else cn_font
    bold_font = "Helvetica-Bold" if lang == "en" else cn_font

    styles = getSampleStyleSheet()
    return {
        "base_font": base_font,
        "bold_font": bold_font,
        "title": ParagraphStyle(
            "QuoteTitle",
            parent=styles["Title"],
            fontName=bold_font,
            fontSize=16,
            leading=20,
            textColor=colors.HexColor(BRAND_BLUE),
            alignment=TA_RIGHT,
            spaceAfter=0,
        ),
        "subtitle": ParagraphStyle(
            "QuoteSubtitle",
            parent=styles["Normal"],
            fontName=base_font,
            fontSize=8.0,
            leading=10,
            textColor=colors.HexColor(TEXT_MUTED),
            alignment=TA_RIGHT,
        ),
        "normal": ParagraphStyle(
            "QuoteNormal",
            parent=styles["Normal"],
            fontName=base_font,
            fontSize=7.2,
            leading=9.6,
            alignment=TA_LEFT,
            textColor=colors.HexColor(TEXT_DARK),
            wordWrap="CJK" if lang == "zh" else None,
        ),
        "small": ParagraphStyle(
            "QuoteSmall",
            parent=styles["Normal"],
            fontName=base_font,
            fontSize=6.8,
            leading=9.0,
            textColor=colors.HexColor(TEXT_MUTED),
            wordWrap="CJK" if lang == "zh" else None,
        ),
        "product_name": ParagraphStyle(
            "ProductName",
            parent=styles["Normal"],
            fontName=bold_font,
            fontSize=8.1,
            leading=10.2,
            textColor=colors.HexColor(BRAND_DARK),
            wordWrap="CJK" if lang == "zh" else None,
        ),
        "normal_cjk": ParagraphStyle(
            "QuoteNormalCJK",
            parent=styles["Normal"],
            fontName=cn_font,
            fontSize=7.2,
            leading=9.6,
            alignment=TA_LEFT,
            textColor=colors.HexColor(TEXT_DARK),
            wordWrap="CJK",
        ),
        "small_cjk": ParagraphStyle(
            "QuoteSmallCJK",
            parent=styles["Normal"],
            fontName=cn_font,
            fontSize=6.8,
            leading=9.0,
            textColor=colors.HexColor(TEXT_MUTED),
            wordWrap="CJK",
        ),
        "product_name_cjk": ParagraphStyle(
            "ProductNameCJK",
            parent=styles["Normal"],
            fontName=cn_font,
            fontSize=8.1,
            leading=10.2,
            textColor=colors.HexColor(BRAND_DARK),
            wordWrap="CJK",
        ),
        "right": ParagraphStyle(
            "QuoteRight",
            parent=styles["Normal"],
            fontName=base_font,
            fontSize=7.0,
            leading=9.2,
            alignment=TA_RIGHT,
            textColor=colors.HexColor(TEXT_DARK),
        ),
        "center": ParagraphStyle(
            "QuoteCenter",
            parent=styles["Normal"],
            fontName=base_font,
            fontSize=7.0,
            leading=9.2,
            alignment=TA_CENTER,
            textColor=colors.HexColor(TEXT_DARK),
        ),
        "header": ParagraphStyle(
            "QuoteTableHeader",
            parent=styles["Normal"],
            fontName=bold_font,
            fontSize=7.1,
            leading=8.6,
            textColor=colors.white,
            alignment=TA_CENTER,
            wordWrap="CJK" if lang == "zh" else None,
        ),
        "section": ParagraphStyle(
            "QuoteSection",
            parent=styles["Heading3"],
            fontName=bold_font,
            fontSize=8.8,
            leading=11,
            textColor=colors.HexColor(BRAND_BLUE),
            spaceBefore=4,
            spaceAfter=3,
        ),
    }


def _draw_footer(canvas, doc, lang: str, base_font: str):
    """Simple ASCII footer to avoid mixed-font spacing issues in generated PDFs."""
    width, _ = A4
    canvas.saveState()
    canvas.setFont("Helvetica", 6.5)
    canvas.setStrokeColor(colors.HexColor(BORDER_BLUE))
    canvas.line(doc.leftMargin, 9 * mm, width - doc.rightMargin, 9 * mm)
    canvas.setFillColor(colors.HexColor(TEXT_MUTED))
    canvas.drawString(doc.leftMargin, 5.5 * mm, "LESSO Plumbing & Sanitary Ware")
    canvas.drawRightString(width - doc.rightMargin, 5.5 * mm, f"Page {doc.page}")
    canvas.restoreState()


def _summary_box(items: pd.DataFrame, lang: str, labels: dict[str, str], styles: dict[str, ParagraphStyle]) -> Table:
    sku_count = int(items["sap"].nunique()) if not items.empty else 0
    total_qty = int(items["quantity"].sum()) if not items.empty else 0
    total_amount = float(items["amount"].sum()) if not items.empty else 0
    total_volume = float(items["total_volume"].sum()) if not items.empty else 0
    data = [[
        _pdf_para(labels["sku"], styles["small"]), _pdf_para(str(sku_count), styles["center"]),
        _pdf_para(labels["qty"], styles["small"]), _pdf_para(str(total_qty), styles["center"]),
        _pdf_para(labels["amount"], styles["small"]), _pdf_para(f"{total_amount:.2f}", styles["right"]),
        _pdf_para(labels["cbm"], styles["small"]), _pdf_para(f"{total_volume:.4f}", styles["right"]),
    ]]
    table = Table(data, colWidths=[17*mm, 16*mm, 18*mm, 16*mm, 27*mm, 27*mm, 18*mm, 22*mm], hAlign="RIGHT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(LIGHT_BG)),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(BORDER_BLUE)),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor(BORDER_BLUE)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    return table


def _product_info_para(row: pd.Series, lang: str, styles: dict[str, ParagraphStyle]) -> Table:
    name = _name(row, lang)
    if lang == "en":
        lines = [
            f"Model: {row.get('model', '')}",
            f"SAP: {row.get('sap', '')}",
            f"Category: {row.get('category', '')}",
        ]
    else:
        lines = [
            f"型号：{row.get('model', '')}",
            f"SAP号：{row.get('sap', '')}",
            f"分类：{row.get('category', '')}",
        ]
    data = [[_pdf_para_auto(name, styles, "product_name")], [_pdf_lines_auto(lines, styles, "small")]]
    t = Table(data, colWidths=[41 * mm])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))
    return t


def export_quote_pdf(items: pd.DataFrame, customer: str, quote_no: str, price_note: str = "", lang: str = "zh") -> bytes:
    """Export an A4 portrait PDF quotation with a more polished business layout. lang='zh' or 'en'."""
    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        rightMargin=9 * mm,
        leftMargin=9 * mm,
        topMargin=9 * mm,
        bottomMargin=13 * mm,
    )

    styles = _make_pdf_styles(lang)
    labels = _field_labels(lang)
    elements: list[object] = []

    # Header / letterhead
    if lang == "en":
        brand_main = "LESSO"
        brand_sub = "Plumbing & Sanitary Ware | International Quotation"
        quote_title = "QUOTATION"
    else:
        # Keep brand Latin and Chinese text in separate fields to avoid mixed-font character crowding.
        brand_main = "LESSO"
        brand_sub = "联塑水暖卫浴 / 外贸报价单"
        quote_title = "正式报价单"

    header_data = [[
        Paragraph(_safe_text(brand_main), ParagraphStyle(
            "BrandMain", fontName="Helvetica-Bold", fontSize=20, leading=23,
            textColor=colors.white, alignment=TA_LEFT
        )),
        Paragraph(_safe_text(quote_title), styles["title"]),
    ], [
        Paragraph(_safe_text(brand_sub), ParagraphStyle(
            "BrandSub", fontName=styles["base_font"], fontSize=8.4, leading=11.2,
            textColor=colors.HexColor("#EEF4FA"), alignment=TA_LEFT,
            wordWrap="CJK" if lang == "zh" else None,
        )),
        Paragraph(_safe_text(datetime.now().strftime("%Y-%m-%d")), styles["subtitle"]),
    ]]
    header = Table(header_data, colWidths=[80 * mm, 112 * mm], hAlign="LEFT")
    header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(BRAND_BLUE)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(header)
    elements.append(Spacer(1, 5))

    # Quotation metadata block
    info_data = [[
        _pdf_para(labels["customer"], styles["small"]), _pdf_para(customer or "", styles["normal"]),
        _pdf_para(labels["quote_no"], styles["small"]), _pdf_para(quote_no, styles["normal"]),
        _pdf_para(labels["date"], styles["small"]), _pdf_para(datetime.now().strftime("%Y-%m-%d"), styles["normal"]),
    ], [
        _pdf_para(labels["rule"], styles["small"]), _pdf_para(price_note, styles["normal"]),
        _pdf_para(labels["currency"], styles["small"]), _pdf_para("USD", styles["normal"]),
        _pdf_para(labels["validity"], styles["small"]), _pdf_para("Subject to final PI" if lang == "en" else "以最终PI为准", styles["normal"]),
    ]]
    info_table = Table(info_data, colWidths=[18*mm, 46*mm, 20*mm, 42*mm, 16*mm, 50*mm], hAlign="LEFT")
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(LIGHT_BG)),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(BORDER_BLUE)),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor(BORDER_BLUE)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 6))

    # Product table - compact but less mechanical than the old raw data table.
    if lang == "en":
        table_header = ["No.", "Image", "Product Information", "Description / Package", "Unit Price\n(USD)", "Qty", "Amount\n(USD)", "CBM"]
    else:
        table_header = ["序号", "图片", "产品信息", "描述 / 包装", "单价\n(USD)", "数量", "金额\n(USD)", "CBM"]

    data: list[list[object]] = [[Paragraph(_safe_text(h).replace("\n", "<br/>"), styles["header"]) for h in table_header]]
    for idx, (_, r) in enumerate(items.iterrows(), start=1):
        img_cell: object = _pdf_para("No image" if lang == "en" else "无图片", styles["small"])
        image_url = r.get("image_url")
        if image_url:
            try:
                pdf_img = _pdf_image_from_url(str(image_url))
                if pdf_img is not None:
                    img_cell = pdf_img
                else:
                    img_cell = _pdf_para("Image error" if lang == "en" else "图片错误", styles["small"])
            except Exception:
                img_cell = _pdf_para("Image error" if lang == "en" else "图片错误", styles["small"])

        package = str(r.get("package_info") or "")
        description = str(r.get("description") or "")
        desc_lines = [description]
        if package:
            desc_lines.append(("Package: " if lang == "en" else "包装：") + package)

        data.append([
            _pdf_para(str(idx), styles["center"]),
            img_cell,
            _product_info_para(r, lang, styles),
            _pdf_lines_auto(desc_lines, styles, "normal"),
            _pdf_para(f"{float(r.get('quote_price', 0)):.2f}", styles["right"]),
            _pdf_para(str(int(r.get("quantity", 1))), styles["center"]),
            _pdf_para(f"{float(r.get('amount', 0)):.2f}", styles["right"]),
            _pdf_para(f"{float(r.get('total_volume', 0)):.4f}", styles["right"]),
        ])

    col_widths = [8*mm, 25*mm, 44*mm, 54*mm, 19*mm, 11*mm, 19*mm, 12*mm]
    table = Table(data, repeatRows=1, colWidths=col_widths, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_BLUE)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), styles["bold_font"]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(BORDER_BLUE)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (0, 1), (1, -1), "CENTER"),
        ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(SOFT_BG)]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 7))

    elements.append(_summary_box(items, lang, labels, styles))
    elements.append(Spacer(1, 7))

    if lang == "en":
        remarks_lines = [
            "1. Product pictures are for reference; final goods are subject to confirmed samples or PI.",
            "2. Prices are generated by the selected pricing rule and should be confirmed before order placement.",
            "3. Stock and volume data are for quotation reference and may change before shipment.",
        ]
        signature_left = "Prepared / Sales"
        signature_right = "Customer Confirmation"
    else:
        remarks_lines = [
            "1. 产品图片仅供参考，最终以确认样品或 PI 为准。",
            "2. 报价按当前所选价格规则自动生成，下单前请再次确认。",
            "3. 库存与体积数据用于报价参考，实际出货前可能调整。",
        ]
        signature_left = "制单 / 业务员"
        signature_right = "客户确认"

    remarks_table = Table([
        [_pdf_para(labels["terms"], styles["product_name"]), _pdf_lines_auto(remarks_lines, styles, "small")],
        [_pdf_para(signature_left, styles["small"]), _pdf_para(signature_right, styles["small"])],
        ["", ""],
    ], colWidths=[35*mm, 157*mm], hAlign="LEFT")
    remarks_table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(BORDER_BLUE)),
        ("INNERGRID", (0, 0), (-1, -1), 0.2, colors.HexColor(BORDER_BLUE)),
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor(LIGHT_BG)),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LINEABOVE", (0, 2), (-1, 2), 0.2, colors.HexColor(BORDER_BLUE)),
        ("ROWHEIGHT", (0, 2), (-1, 2), 13*mm),
    ]))
    elements.append(KeepTogether(remarks_table))

    doc.build(
        elements,
        onFirstPage=lambda canvas, d: _draw_footer(canvas, d, lang, styles["base_font"]),
        onLaterPages=lambda canvas, d: _draw_footer(canvas, d, lang, styles["base_font"]),
    )
    return bio.getvalue()
